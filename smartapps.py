
import streamlit as st
from openpyxl import load_workbook, Workbook
from pathlib import Path
from datetime import datetime

st.title("チェックリストアプリ（スマホ対応）")

items = ["電源投入エラーなし", "温度センサー正常", "ジャム履歴なし"]
checked = [item for item in items if st.checkbox(item)]

# --- ここを変更 ---
# 例：C:/Users/<ユーザー名>/OneDrive - Advantest/checklist.xlsx に保存
one_drive_dir = Path.home() / "OneDrive - Advantest"
target_path = one_drive_dir / "checklist.xlsx"

if st.button("結果を保存"):
    try:
        target_path.parent.mkdir(parents=True, exist_ok=True)

        if not target_path.exists():
            wb = Workbook()
            wb.active.title = "Checklist"
            wb.save(target_path)

        wb = load_workbook(target_path)
        ws = wb.active
        start_row = ws.max_row + 1
        ws.cell(row=start_row, column=1, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=start_row, column=2, value="✔".join(checked) if checked else "")
        wb.save(target_path)

        st.success(f"保存しました → {target_path}")

    except PermissionError as e:
        st.error("ファイルが他のアプリで開かれている可能性があります。Excelや同期中のOneDriveを一度閉じてから再試行してください。")
        st.code(str(e))
    except Exception as e:
        st.error("保存でエラーが発生しました。パスや権限をご確認ください。")
        st.code(str(e))

    
    
    
